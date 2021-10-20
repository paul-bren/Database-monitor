import pyodbc
from pyodbc import Error
import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Series, Reference
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.descriptors.serialisable import Serialisable
import azutil

server = input("Please enter server: ")
database = input("Please enter db: ")
username = azutil.username
password = azutil.password
driver= "{ODBC Driver 13 for SQL Server}"

def server_details(conn, query, writer, sheet_name):
    cursor = conn.cursor
    data = pd.read_sql(query, conn)
    data.to_excel(writer, sheet_name, startrow=2, startcol=1, index = False, header = False)

def all_db_details(conn, query, writer, sheet_name):
    cursor = conn.cursor
    data = pd.read_sql(query, conn)
    data.to_excel(writer, sheet_name, startrow=writer.sheets['Overview'].max_row, startcol=1, index = False, header = False)
    
def top_dtu_queries(conn, query):
    query_id_list = []
    cursor = conn.cursor()
    data = cursor.execute(query)
    for row in data:
        query_id_list.append(row[0])
    return query_id_list


def db_overview(writer, sheet_name):
    data1 = []
    data2 = pd.DataFrame(data1)
    data2.to_excel(writer, sheet_name, startrow=2, startcol=0, index = False)

def blocking_sessions_now(conn, query, writer, sheet_name):
    cursor = conn.cursor
    data = pd.read_sql(query, conn)
    data.to_excel(writer, sheet_name, startrow=2, startcol=0, index = False)

def blocked_sessions_now(conn, query, writer, sheet_name):
    cursor = conn.cursor()
    data = pd.read_sql(query, conn)
    data.to_excel(writer, sheet_name, startrow=2, startcol=4)

def large_tables(conn, query, writer, sheet_name):
    cursor = conn.cursor()
    data = pd.read_sql(query, conn)
    data.to_excel(writer, sheet_name, startrow=2, startcol=0, index = False)

def large_table_format(ws, chart1, data_list, data_names):
    chart1.add_data(data_list, titles_from_data=True)
    chart1.set_categories(data_names)
    chart1.shape = 4
    chart1.height = 13
    chart1.width = 28

def sql_query_text(conn, query, query_id):
    cursor = conn.cursor()
    data = pd.read_sql(query, conn, params=[query_id])
    return data
    
def find_table_name_from_query_text(conn, query, all_tables, query_id):
    table_list =[]
    final_table_list = []
    cursor = conn.cursor()
    data = pd.read_sql(query, conn, params=[query_id])
    data2 = pd.read_sql(all_tables, conn)
    for query_string in data['query_text']:
        word = query_string.split(' ')
    for table in data2['name']:
        table_list.append(table)
    for found_tables in table_list:
        if found_tables in word:
            final_table_list.append(found_tables)
        else:
            continue
    return final_table_list

def stale_statistics(conn, query, table, writer, sheet_name, row, column_header):
    cursor = conn.cursor()
    data = pd.read_sql(query, conn, params=[table])
    data.to_excel(writer, sheet_name, startrow=row, startcol=0, index = False, header = column_header)

def index_stats(conn, query, table, writer, sheet_name, row, column_header):
    cursor = conn.cursor()
    data = pd.read_sql(query, conn, params=[table])
    data.to_excel(writer, sheet_name, startrow=row, startcol=0, index = False, header = column_header)
 

def dataframe_size(conn, query, table, df_size):
    cursor = conn.cursor()
    data = pd.read_sql(query, conn, params=[table])
    df_size = len(data)
    return df_size


def curr_DTU(conn, query, writer, sheet_name):
    cursor = conn.cursor()
    data = pd.read_sql(query, conn)
    data.to_excel(writer, sheet_name, startrow=2, startcol=0, index = False)

def dtu_val(conn, query, writer, sheet_name):
    cursor = conn.cursor()
    data = pd.read_sql(query, conn)
    data.to_excel(writer, sheet_name, startrow=9, startcol=0, index = False)

def price_tier(conn, query, writer, sheet_name):
    cursor = conn.cursor()
    data = pd.read_sql(query, conn)
    data.to_excel(writer, sheet_name, startrow=14, startcol=0, index = False)

def database_size(conn, query, writer, sheet_name):
    cursor = conn.cursor()
    data = pd.read_sql(query, conn)
    data.to_excel(writer, sheet_name, startrow=19, startcol=0, index = False)

def current_frag(conn, query, table, writer, sheet_name, row, column_header):
    cursor = conn.cursor()
    data = pd.read_sql(query, conn, params=[table])
    data.to_excel(writer, sheet_name, startrow=row, startcol=0, index = False, header = column_header)

def unused_index(conn, query, writer, sheet_name):
    cursor = conn.cursor()
    data = pd.read_sql(query, conn)
    data.to_excel(writer, sheet_name, startrow=2, startcol=0, index = False)

def largest_admin_indexes(conn, query, writer, sheet_name):
    cursor = conn.cursor()
    data = pd.read_sql(query, conn)
    data.to_excel(writer, sheet_name, startrow=20, startcol=0, index = False)

def maxdop_val(conn, query, writer, sheet_name):
    cursor = conn.cursor()
    data = pd.read_sql(query, conn)
    data.to_excel(writer, sheet_name, startrow=2, startcol=0, index = False)

top_dtu_sql = (''' SELECT TOP 3
q.query_id, 
convert(decimal(18, 2), SUM(count_executions * avg_cpu_time / 1000.0 / 1000.0)) AS total_cpu_sec, 
qt.query_sql_text
from sys.query_store_query q
join sys.query_store_plan p on q.query_id = p.query_id
join sys.query_store_runtime_stats rs on rs.plan_id=p.plan_id
join sys.query_store_runtime_stats_interval rsi on  rsi.runtime_stats_interval_id=rs.runtime_stats_interval_id
join sys.query_store_query_text qt on q.query_text_id = qt.query_text_id
WHERE rsi.start_time >= DATEADD(MINUTE, -3000, GETUTCDATE())
AND rsi.start_time <= DATEADD(MINUTE, -370, GETUTCDATE())
-- WHERE rsi.start_time >= convert(varchar,'2019-10-31 08:00:00',120) and rsi.start_time < convert(varchar,'2019-10-31 09:00:00',120)
group by q.query_id,qt.query_sql_text,rsi.start_time,rsi.end_time
order by total_cpu_sec desc
''')
server_name = (''' SELECT CONCAT(CONVERT(nvarchar(50),SERVERPROPERTY('ServerName')),'.database.windows.net') as FullServerName ''')
database_name = (''' SELECT DB_NAME() AS [Current Database]''')
#See following URL for reason behind converting next query to have NVARCHAR: https://stackoverflow.com/questions/45749323/pandas-error-when-creating-dataframe-from-ms-sql-server-database-odbc-sql-type?rq=1
database_coll = (""" SELECT CAST(SERVERPROPERTY('Collation') as NVARCHAR(4000)) """)
database_timezone = (''' SELECT CURRENT_TIMEZONE()''')
database_query_store_val = (''' SELECT actual_state_desc
FROM sys.database_query_store_options''')
database_query_store_cur_size = ('''SELECT  current_storage_size_mb
FROM sys.database_query_store_options ''')
database_query_store_max_size = ('''SELECT  max_storage_size_mb
FROM sys.database_query_store_options; ''')
all_tables = (''' select name from sys.tables ''')
    
largest_tables = ('''select top 10 schema_name(tab.schema_id) + '.' + tab.name as [table],  --change 15 to the amount of tables you want to list
    cast(sum(spc.used_pages * 8)/1024.00 as numeric(36, 2)) as used_mb,
    cast(sum(spc.total_pages * 8)/1024.00 as numeric(36, 2)) as allocated_mb
    from sys.tables tab
    join sys.indexes ind 
     on tab.object_id = ind.object_id
    join sys.partitions part 
     on ind.object_id = part.object_id and ind.index_id = part.index_id
    join sys.allocation_units spc
     on part.partition_id = spc.container_id
    group by schema_name(tab.schema_id) + '.' + tab.name
    order by sum(spc.used_pages) desc;''')
sql_text = (''' SELECT qt.query_sql_text query_text
 FROM sys.query_store_query q
 JOIN sys.query_store_query_text qt ON q.query_text_id = qt.query_text_id
 WHERE q.query_id = ? ''')
stale_stats = ('''SELECT TOP 50
     CONCAT(sch.name, '.', obj.name) AS 'Table Name'
    ,MAX(sp.last_updated)           AS 'Stats Last Updated'
    ,MAX(sp.rows)                   AS 'Rows'
    ,MAX(sp.modification_counter)   AS 'Modification Counter'
    FROM sys.stats st
    JOIN sys.objects obj ON st.object_id = obj.object_id
    JOIN sys.schemas sch ON obj.schema_id = sch.schema_id
    CROSS APPLY sys.dm_db_stats_properties(obj.object_id, st.stats_id) AS sp
    WHERE obj.is_ms_shipped = 0
    AND obj.name = ?
    GROUP BY CONCAT(sch.name, '.', obj.name)
    ORDER BY MAX(sp.modification_counter) DESC''')
index_stats_query = ('''
SET NOCOUNT ON

declare @idxIdentifierBegin char(1), @idxIdentifierEnd char(1);
declare @statsIdentifierBegin char(1), @statsIdentifierEnd char(1);
declare @TableUsedInQuery nvarchar(200)

--Step 1
--review query with perfromance problem and list all the tables used by query
--add table name to @TableUsedInQuery

set @TableUsedInQuery= ?


drop table if exists #statsBefore
drop table if exists #IndexStats

SELECT OBJECT_NAME(IX.OBJECT_ID) Table_Name
	   ,IX.name AS Index_Name
	   ,IX.type_desc Index_Type
	   ,SUM(PS.[used_page_count]) * 8 IndexSizeKB
	   ,IXUS.user_seeks AS NumOfSeeks
	   ,IXUS.user_scans AS NumOfScans
	   ,IXUS.user_lookups AS NumOfLookups
	   ,IXUS.user_updates AS NumOfUpdates
	   ,IXUS.last_user_seek AS LastSeek
	   ,IXUS.last_user_scan AS LastScan
	   ,IXUS.last_user_lookup AS LastLookup
	   ,IXUS.last_user_update AS LastUpdate
into #IndexStats

FROM sys.indexes IX
INNER JOIN sys.dm_db_index_usage_stats IXUS ON IXUS.index_id = IX.index_id AND IXUS.OBJECT_ID = IX.OBJECT_ID
INNER JOIN sys.dm_db_partition_stats PS on PS.object_id=IX.object_id  and ps.[index_id] = ix.[index_id]
WHERE OBJECT_NAME(IX.OBJECT_ID) = @TableUsedInQuery
GROUP BY OBJECT_NAME(IX.OBJECT_ID) ,IX.name ,IX.type_desc ,IXUS.user_seeks ,IXUS.user_scans ,IXUS.user_lookups,IXUS.user_updates ,IXUS.last_user_seek ,IXUS.last_user_scan ,IXUS.last_user_lookup ,IXUS.last_user_update


		select 
			ObjectSchema = OBJECT_SCHEMA_NAME(s.object_id)
			,ObjectName = object_name(s.object_id) 
			,s.object_id
			,s.stats_id
			,StatsName = s.name
			,sp.last_updated
			,sp.rows
			,sp.rows_sampled
			,sp.modification_counter
			, i.type
			, i.type_desc
			,0 as SkipStatistics
		into #statsBefore
		from sys.stats s cross apply sys.dm_db_stats_properties(s.object_id,s.stats_id) sp 
		left join sys.indexes i on sp.object_id = i.object_id and sp.stats_id = i.index_id
		where OBJECT_SCHEMA_NAME(s.object_id) != 'sys' and /*Modified stats or Dummy mode*/(isnull(sp.modification_counter,0)>=0 )--or @mode='dummy')
		and 
		s.object_id = OBJECT_ID(@TableUsedInQuery)
		order by sp.last_updated asc

	if exists(
			select 1
			from #statsBefore 
			where StatsName like '%[%' or StatsName like '%]%'
			or ObjectSchema like '%[%' or ObjectSchema like '%]%'
			or ObjectName like '%[%' or ObjectName like '%]%'
			)
		begin
			set @statsIdentifierBegin = '"'
			set @statsIdentifierEnd = '"'
		end
		else 
		begin
			set @statsIdentifierBegin = '['
			set @statsIdentifierEnd = ']'
		end

				select StatsName as Index_name
		, 
		last_updated,
		CurrentStatInfo = '#rows:' + cast([rows] as varchar(100)) + ' #modifications:' + cast(modification_counter as varchar(100)) + ' modification percent: ' + format((1.0 * modification_counter/ rows ),'p')
		from #statsBefore
		WHERE StatsName like lower('IDX%')
		order by modification_counter desc

		--select ''+ @statsIdentifierBegin + ObjectSchema + +@statsIdentifierEnd + '.'+@statsIdentifierBegin + ObjectName + @statsIdentifierEnd +' (' + @statsIdentifierBegin + StatsName + @statsIdentifierEnd + ')' as Index_name

''')
current_DTU = ('''SELECT end_time,avg_cpu_percent, avg_data_io_percent, avg_log_write_percent
FROM sys.dm_db_resource_stats
ORDER BY end_time DESC;''')
dtu_value = ('''select top(1) dtu_limit from sys.dm_db_resource_stats order by end_time DESC ''')
pricing_tier=(''' select edition from sys.database_service_objectives ''')
service_objective = (''' select service_objective from sys.database_service_objectives ''')
database_size = (''' select convert(decimal(18, 2), SUM(CAST(FILEPROPERTY(name, 'SpaceUsed') AS int)/128.0)) AS Used_Space_MB
FROM sys.database_files
GROUP BY type_desc
HAVING type_desc = 'ROWS' ''')
current_frag_temp = ('''SELECT
T.name as 'Table',
I.name as 'Index',
DDIPS.avg_fragmentation_in_percent,
DDIPS.page_count
FROM sys.dm_db_index_physical_stats (DB_ID(), NULL, NULL, NULL, NULL) AS DDIPS
INNER JOIN sys.tables T on T.object_id = DDIPS.object_id
INNER JOIN sys.schemas S on T.schema_id = S.schema_id
INNER JOIN sys.indexes I ON I.object_id = DDIPS.object_id
AND DDIPS.index_id = I.index_id
WHERE DDIPS.database_id = DB_ID()
and T.name = ?
and I.name like (lower('IDX%'))
and I.name is not null
AND DDIPS.avg_fragmentation_in_percent >= 0
ORDER BY DDIPS.avg_fragmentation_in_percent desc ''')
unused_indexes = ('''SELECT TOP 25
o.name AS ObjectName
, i.name AS IndexName
, i.index_id AS IndexID
, dm_ius.user_seeks AS UserSeek
, dm_ius.user_scans AS UserScans
, dm_ius.user_lookups AS UserLookups
, dm_ius.user_updates AS UserUpdates
, p.TableRows
, 'DROP INDEX ' + QUOTENAME(i.name)
+ ' ON ' + QUOTENAME(s.name) + '.'
+ QUOTENAME(OBJECT_NAME(dm_ius.OBJECT_ID)) AS 'drop statement'
FROM sys.dm_db_index_usage_stats dm_ius
INNER JOIN sys.indexes i ON i.index_id = dm_ius.index_id 
AND dm_ius.OBJECT_ID = i.OBJECT_ID
INNER JOIN sys.objects o ON dm_ius.OBJECT_ID = o.OBJECT_ID
INNER JOIN sys.schemas s ON o.schema_id = s.schema_id
INNER JOIN (SELECT SUM(p.rows) TableRows, p.index_id, p.OBJECT_ID
FROM sys.partitions p GROUP BY p.index_id, p.OBJECT_ID) p
ON p.index_id = dm_ius.index_id AND dm_ius.OBJECT_ID = p.OBJECT_ID
WHERE OBJECTPROPERTY(dm_ius.OBJECT_ID,'IsUserTable') = 1
AND dm_ius.database_id = DB_ID()
AND i.type_desc = 'nonclustered'
AND i.is_primary_key = 0
AND i.is_unique_constraint = 0
and dm_ius.user_seeks = 0
and dm_ius.user_scans = 0
and dm_ius.user_lookups = 0
ORDER BY dm_ius.user_updates desc ''')
largest_dba_indexes = (''' SELECT TOP(10)
OBJECT_SCHEMA_NAME(i.OBJECT_ID) AS SchemaName,
OBJECT_NAME(i.OBJECT_ID) AS TableName,
i.name AS IndexName,
8 * SUM(a.used_pages) AS 'Indexsize(MB)'
FROM sys.indexes AS i
JOIN sys.partitions AS p ON p.OBJECT_ID = i.OBJECT_ID AND p.index_id = i.index_id
JOIN sys.allocation_units AS a ON a.container_id = p.partition_id
WHERE i.name like 'IDX%'
GROUP BY i.OBJECT_ID,i.index_id,i.name
--ORDER BY OBJECT_NAME(i.OBJECT_ID),i.index_id
ORDER BY 'Indexsize(MB)' DESC ''')
current_maxdop_value = ('''SELECT CAST([value] as nvarchar(100)) as 'MAXDOP_VALUE' FROM sys.database_scoped_configurations WHERE [name] = 'MAXDOP'; ''')
blocking_time_threshold = 900000
blocking_sessions = (''' select A.blocking_session_id, C.status, st.text as blocking_sess_last_query
FROM
(SELECT blocking_session_id, max(wait_time) as longest_wait_time
FROM sys.dm_exec_requests
WHERE  blocking_session_id <> 0
AND blocking_session_id not in (select session_id from sys.dm_exec_requests where blocking_session_id <> 0)
group by blocking_session_id) A
LEFT JOIN sys.dm_exec_requests B ON A.blocking_session_id = B.session_id
JOIN sys.dm_exec_sessions C on A.blocking_session_id = C.session_id
JOIN sys.dm_exec_connections D on D.session_id = C.session_id
OUTER APPLY sys.dm_exec_sql_text(D.most_recent_sql_handle) st
WHERE A.longest_wait_time > ''' + str(blocking_time_threshold))

blocked_sessions = (""" SELECT  req.Session_id,
        req.blocking_session_id,
	    req.wait_time/1000 wait_time_in_seconds,
	    req.wait_type,
        SUBSTRING(ST.text, (req.statement_start_offset / 2)+1, ((CASE statement_end_offset WHEN -1 THEN DATALENGTH(ST.text)ELSE req.statement_end_offset END-req.statement_start_offset)/ 2)+1) AS query_text
FROM    sys.dm_exec_requests req        
        OUTER APPLY sys.dm_exec_sql_text(req.sql_handle) st
WHERE   req.blocking_session_id <> 0
AND req.wait_time > """ + str(blocking_time_threshold))

def colour(mycell):
    redFill = PatternFill(start_color='0000FF00',end_color='0000FF00',fill_type='solid')
    mycell.font = Font(size=18)
    mycell.fill = redFill
    

def main():
    try:
        conn = azutil.connect_db(server, database)
        print ("Connection successfull...")
    except: ("Connection to database failed")
    #I will need to identify a proper temp location to create excel file. Then the identified directory will need to be regularly cleaned up
    writer = pd.ExcelWriter("location", engine = 'openpyxl')
    sheet_name = 'Overview'
    db_overview(writer, sheet_name)
    server_details(conn, server_name, writer, sheet_name)
    all_db_details(conn, database_name, writer, sheet_name)
    all_db_details(conn, database_size, writer, sheet_name)
    all_db_details(conn, pricing_tier, writer, sheet_name)
    all_db_details(conn, service_objective, writer, sheet_name)
    all_db_details(conn, database_coll, writer, sheet_name)
    all_db_details(conn, database_timezone, writer, sheet_name)
    all_db_details(conn, database_query_store_val, writer, sheet_name)
    all_db_details(conn, database_query_store_cur_size, writer, sheet_name)
    all_db_details(conn, database_query_store_max_size, writer, sheet_name)
    all_db_details(conn, dtu_value, writer, sheet_name)
    all_db_details(conn, current_maxdop_value, writer, sheet_name)
    
    sheet_name = 'Large Tables'
    large_tables(conn, largest_tables, writer, sheet_name)
    top_dtu_query_id = top_dtu_queries(conn, top_dtu_sql)
    sheet_name = 'DTU'
    curr_DTU(conn, current_DTU, writer, sheet_name)
    sheet_name = 'Indexes'
    sheet_name = 'Blocking Sessions'
    blocking_sessions_now(conn, blocking_sessions, writer, sheet_name)
    blocked_sessions_now(conn, blocked_sessions, writer, sheet_name)
    writer.save()
    conn.close()


    path = r""
    wb = load_workbook(path)
    #ws = wb.active -- This selects the default sheet. This won't work if you have more than one sheet.
    # This loop will iterate through all the worksheets
    for sheet in wb.sheetnames:
        name = wb[sheet].title
        ws = wb[sheet]
        if name == "DTU":
            mycell=wb[sheet]['A1']
            mycell.value = "Please find DTU details for the past hour below "
            for row in ws[2:ws.max_row]:
                cell_A = row[:1][0]
                cell_A.alignment = Alignment(horizontal='center')  
            mycell.alignment = Alignment(horizontal='center')
            colour(mycell)
        elif name =="Indexes":
            mycell=wb[sheet]['A1']
            mycell.value = "Please find details for unused " + name + " below "
            colour(mycell)
            mycell=wb[sheet]['A15']
            mycell.value = "Top 10 Largest Indexes"
            colour(mycell)
        elif name == "Large Tables":
            mycell=wb[sheet]['A1']
            mycell.value = "Please find details for " + name + " below "
            colour(mycell)
            chart1 = BarChart()
            chart1.type = "bar"
            chart1.style = 11
            chart1.title = "Top 10 Largest Tables"
            chart1.y_axis.title = 'Size in MB'
            chart1.x_axis.title = 'Table Name'
            data_list = Reference(ws, min_col=2, min_row=ws.min_row+2, max_row=ws.max_row, max_col=3)
            data_names = Reference(ws, min_col=1, min_row=4, max_row=13, max_col=1)
            large_table_format(ws, chart1, data_list, data_names)
            ws.add_chart(chart1, "A" + str(ws.max_row+3))
      

        elif name == "Overview":
            mycell=wb[sheet]['A1']
            mycell.value = "Database overview"
            colour(mycell)
            cust_data = ['Server Name:', 'Database Name:', 'Database Size:', 'Database Pricing Tier:', 'Database Service Objective:', 'Database Collation:', 'Database Timezone:', 'Database Query Store Value:', 'Database Query Store Size:', 'Database Query Store Max Size:', 'DTU Size:', 'MaxDOP Value:'  ]
            count = 0
            counter = 3
            for items in cust_data:
                if count == 0:
                    mycell=wb[sheet]['A3']
                    mycell2=wb[sheet]['B3']
                    mycell.font = Font(size=16, bold = True)
                    mycell.alignment = Alignment(horizontal='left')
                    mycell2.alignment = Alignment(horizontal='center')
                    mycell.value = cust_data[count]
                else:
                    mycell=wb[sheet]['A' + str(counter)]
                    mycell2=wb[sheet]['B' + str(counter)]
                    mycell.font = Font(size=16, bold = True)
                    mycell.alignment = Alignment(horizontal='left')
                    mycell2.alignment = Alignment(horizontal='center')
                    mycell.value = cust_data[count]
                count +=1
                counter +=1
            mycell=wb[sheet]['B3']
            mycell.alignment = Alignment(horizontal='center')
     
        else:
            mycell=wb[sheet]['A1']
            mycell.value = "Please find details for " + name + " below "
            colour(mycell)
        #The purpose of this loop is to iterate through all the columns in each of the worksheets and expand the column length when the workbook is opened
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try: # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    print("Error completing remaining sheets")
                    pass
            adjusted_width = (max_length + 4) * 1.4
            ws.column_dimensions[column].width = adjusted_width #This sets the width.
 
        wb.save(path)
        wb.close
        
    
        
if __name__ == "__main__":
    main()



